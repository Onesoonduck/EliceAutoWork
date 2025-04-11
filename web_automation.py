from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from tkinter import Tk
from tkinter.filedialog import askopenfilename


import time
import os
import glob
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import os

from config import LOGIN_INFO, EXCEL_SETTINGS


class WebAutomation:
    def __init__(self):
        self.driver = None
        self.setup_driver()

    def run_full_report_process(self):
        try:
            url = input("▶ 이동할 대시보드 URL을 입력하세요:\n> ").strip()
            chapter_num = input("▶ 챕터 번호를 입력하세요:\n> ").strip()

            if not chapter_num.isdigit():
                raise ValueError("챕터 번호는 숫자여야 합니다.")

            self.login()
            self.navigate_to_chapter_report(url, chapter_num)
        finally:
            self.close()

    def setup_driver(self):
        options = webdriver.ChromeOptions()
        prefs = {
            "download.default_directory": os.path.abspath(EXCEL_SETTINGS['download_path']),
            "download.prompt_for_download": False
        }
        options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        self.driver.set_window_size(1920, 1080)
        print("브라우저 창 크기를 1920x1080으로 설정했습니다.")

    def login(self):
        self.driver.get("https://accounts.elice.io/accounts/signin/me?continue_to=https%3A%2F%2Felice.io%2Fuser%2Fprofile&lang=ko")
        time.sleep(2)
        try:
            email_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email']"))
            )
            email_input.send_keys(LOGIN_INFO['username'])
            password_input = self.driver.find_element(By.CSS_SELECTOR, "input[type='password']")
            password_input.send_keys(LOGIN_INFO['password'])
            login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            login_button.click()
            time.sleep(5)
        except Exception as e:
            print(f"로그인 중 오류 발생: {str(e)}")
            raise

    def find_and_click_element(self, by, value, scroll=True):
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((by, value))
            )
            if scroll:
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(1)
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((by, value))
            )
            self.driver.execute_script("arguments[0].click();", element)
            return True
        except Exception as e:
            print(f"요소 찾기/클릭 실패: {str(e)}")
            return False

    def navigate_to_chapter_report(self, url, chapter_num):
        try:
            self.driver.get(url)
            time.sleep(5)

            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '목차')]"))
                )
            except:
                if not self.find_and_click_element(By.XPATH, "//button[contains(text(), '목록')]"):
                    raise Exception("목록 버튼을 찾을 수 없습니다.")
                time.sleep(3)
                if not self.find_and_click_element(By.XPATH, "//button[contains(text(), '목차')]"):
                    raise Exception("목차 버튼을 찾을 수 없습니다.")
            time.sleep(3)

            if not self.find_and_click_element(
                By.XPATH, "//div[contains(@class, 'MuiListItemButton-root')]//p[text()='챕터 테스트']"
            ):
                raise Exception("챕터 테스트를 찾을 수 없습니다.")
            time.sleep(3)

            self.find_and_click_element(By.XPATH, "//button[contains(text(), '목록')]")
            time.sleep(3)

            selectors = [
                (By.CSS_SELECTOR, "button.MuiButton-containedInherit[aria-label='']"),
                (By.XPATH, "//button[contains(text(), '과목 리포트')]"),
                (By.CSS_SELECTOR, "button.MuiButton-root")
            ]

            for by, value in selectors:
                if self.find_and_click_element(by, value):
                    time.sleep(5)
                    print("과목 리포트 다운로드 완료")

                    try:
                        latest_file = self.get_latest_downloaded_file()
                        if latest_file:
                            print(f"📁 다운로드된 엑셀 파일: {os.path.basename(latest_file)}")

                            filtered_file = self.save_name_and_progress_only(latest_file, chapter_num)

                            if filtered_file:
                                df = pd.read_excel(filtered_file)
                                name_progress_dict = dict(zip(df["이름"], df["학습진행률"]))

                                print("▶ 업데이트할 엑셀 파일을 선택해주세요.")
                                Tk().withdraw()  # Tk 창 숨기기

                                # 기본 경로를 OneDrive로 설정
                                default_onedrive = os.path.join(Path.home(), "OneDrive")

                                existing_path = askopenfilename(
                                    title="엑셀 파일 선택",
                                    initialdir=default_onedrive,
                                    filetypes=[("Excel files", "*.xlsx *.xls")]
                                )

                                if not existing_path:
                                    print("❌ 파일이 선택되지 않았습니다.")
                                    return


                                part_ranges = detect_part_ranges(existing_path)
                                update_excel_with_progress(existing_path, name_progress_dict, f"CH {chapter_num}", part_ranges)
                                return latest_file
                    except Exception as e:
                        print(f"엑셀 파일 처리 중 오류 발생: {str(e)}")
                    return
            raise Exception("과목 리포트 버튼을 찾을 수 없습니다.")
        except Exception as e:
            print(f"오류 발생: {str(e)}")
            raise

    def get_latest_downloaded_file(self, timeout=30):
        path = os.path.abspath(EXCEL_SETTINGS['download_path'])
        start_time = time.time()
        print("📥 다운로드 완료를 기다리는 중...")

        while True:
            list_of_files = [
                f for f in glob.glob(os.path.join(path, 'report_adtrack*'))
                if not f.endswith(".crdownload")
            ]

            if list_of_files:
                latest_file = max(list_of_files, key=os.path.getctime)
                break

            if time.time() - start_time > timeout:
                print("❌ 다운로드 대기 시간 초과")
                return None

            time.sleep(1)

        if not latest_file.endswith(".xlsx"):
            new_name = latest_file + ".xlsx"
            os.rename(latest_file, new_name)
            latest_file = new_name

        print(f"✅ 다운로드 완료: {latest_file}")
        return latest_file

    def save_name_and_progress_only(self, input_file, chapter_num):
        try:
            df = pd.read_excel(input_file)
            df["학습진행률"] = df["학습진행률"].str.rstrip('%').astype(int)
            selected_df = df[["이름", "학습진행률"]]
            selected_df["챕터"] = chapter_num

            filename = os.path.basename(input_file)
            save_path = os.path.join(os.path.dirname(input_file), f"filtered_{filename}")
            selected_df.to_excel(save_path, index=False)
            print(f"\n✅ 필터링된 엑셀 저장 완료: {save_path}")
            return save_path
        except Exception as e:
            print(f"이름, 학습진행률 저장 중 오류 발생: {str(e)}")
            return None

    def close(self):
        if self.driver:
            self.driver.quit()


def update_excel_with_progress(file_path, name_progress_dict, chapter_title, part_ranges):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 기존 엑셀에서 이름 목록 가져오기
        existing_names = []
        for row in range(4, ws.max_row):
            name = ws.cell(row=row, column=2).value
            if name:
                existing_names.append(name)

        # 챕터 열 찾기
        target_col = None
        for col in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(row=3, column=col).value).strip()
            if cell_val == chapter_title:
                target_col = col
                break

        if not target_col:
            print(f"❌ '{chapter_title}' 열을 찾을 수 없습니다.")
            return

        # 이름별 진도율 입력 + 없으면 삭제 대상
        rows_to_delete = []
        for row in range(4, ws.max_row):
            name = ws.cell(row=row, column=2).value
            if name in name_progress_dict:
                progress = name_progress_dict[name] / 100
                cell = ws.cell(row=row, column=target_col)
                cell.value = progress
                cell.number_format = '0%'
            else:
                rows_to_delete.append(row)

        # 이름 없으면 행 삭제 (역순으로)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # 저장
        wb.save(file_path)
        print(f"✅ 엑셀 진도율 업데이트 완료: {file_path}")

    except Exception as e:
        print(f"❌ 오류 발생: {str(e)}")


def detect_part_ranges(file_path: str, sheet_name: str = None):
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    part_avg_cols = []
    col_ranges = []
    part_ranges = {}
    prev = 2
    header_row = 3

    for col in range(1, ws.max_column + 1):
        cell_value = str(ws.cell(row=header_row, column=col).value).strip()
        if "파트평균" in cell_value:
            part_avg_cols.append(col)

    for i, col in enumerate(part_avg_cols):
        part_name = f"PART{i+1}"
        part_ranges[part_name] = (prev + 1, col - 1, col)
        prev = col

    return part_ranges
